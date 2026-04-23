"""
Source-to-Taxonomy Mapper
==========================
Maps source items (events, issues, findings, etc.) to L2 taxonomy
categories using spaCy semantic similarity.

Each source item can map to multiple L2s when it legitimately spans more
than one category. Raw scores are replaced with plain-language mapping
statuses (Suggested Match / Needs Review / No Match) for reviewer use.

Adapted from the ORE/PRSA/RAP mappers in the risk_taxonomy_transformer
project. Key patterns carried forward:

- Aggregate reference vectors by L2 name, folding L3/L4 text into each
  L2's semantic vector (enterprise taxonomy files are often at L4 grain
  with L2 repeated across rows; building one vector per row produces
  tied top-N rankings and everything falls into Needs Review).
- Auto-tune the ambiguity margin threshold from the data's P25 margin
  (clamped to [0.01, 0.05]). spaCy cosine scores are compressed; tighter
  thresholds than TF-IDF.
- Classify into three statuses with candidate lists that reviewers
  confirm (instead of forcing a single pick on ambiguous cases).
- All column names live in config/mapper_config.yaml — never hardcode.

Usage:
    python mapper.py

Input:
    - data/input/Taxonomy.xlsx
    - data/input/source_*.xlsx
Output:
    - data/output/mapping_{timestamp}.xlsx
      Sheet 1: All Mappings
      Sheet 2: Needs Review
      Sheet 3: Summary
      Sheet 4: L2 Distribution
      Sheet 5: Raw Scores (hidden)
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
import spacy
import yaml

_PROJECT_ROOT = Path(__file__).parent
_LOG_DIR = _PROJECT_ROOT / "logs"
_LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_LOG_DIR / "mapper_log.txt", mode="w"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION (loaded from config/mapper_config.yaml)
# =============================================================================

_CONFIG_PATH = _PROJECT_ROOT / "config" / "mapper_config.yaml"
with open(_CONFIG_PATH, "r", encoding="utf-8") as _f:
    _cfg = yaml.safe_load(_f)

_src_cfg = _cfg.get("source", {})
_tax_cfg = _cfg.get("taxonomy", {})

# --- spaCy + thresholds ---
SPACY_MODEL = _cfg.get("spacy_model", "en_core_web_lg")
MIN_SIMILARITY_SCORE = _cfg.get("min_similarity_score", 0.50)
HIGH_SIMILARITY_SCORE = _cfg.get("high_similarity_score", 0.75)
AMBIGUITY_MARGIN_THRESHOLD = _cfg.get("ambiguity_margin_threshold", None)

# --- Source file ---
SOURCE_FILE_PATTERN = _src_cfg.get("file_pattern", "source_*.xlsx")
SOURCE_ID_COL = _src_cfg.get("id", "Event ID")
SOURCE_ENTITY_COL = _src_cfg.get("entity_id", "Audit Entity ID")
SOURCE_TITLE_COL = _src_cfg.get("title", "Event Title")
SOURCE_DESC_COL = _src_cfg.get("description", "Event Description / Summary")
SOURCE_CLASS_COL = _src_cfg.get("classification", "Final Event Classification")
SOURCE_STATUS_COL = _src_cfg.get("status", "Event Status")

# --- Taxonomy file ---
TAXONOMY_FILE = _tax_cfg.get("file_name", "Taxonomy.xlsx")
TAX_L1_COL = _tax_cfg.get("l1", "L1")
TAX_L2_COL = _tax_cfg.get("l2", "L2")
TAX_L2_DEF_COL = _tax_cfg.get("l2_definition", "L2 Definition")
TAX_L3_COL = _tax_cfg.get("l3", "L3")
TAX_L3_DEF_COL = _tax_cfg.get("l3_definition", "L3 Definition")
TAX_L4_COL = _tax_cfg.get("l4", "L4")
TAX_L4_DEF_COL = _tax_cfg.get("l4_definition", "L4 Definition")

# --- Status filter ---
CLOSED_STATUSES = {s.lower() for s in _cfg.get("closed_statuses", [])}


# =============================================================================
# CORE FUNCTIONS
# =============================================================================

def load_taxonomy(input_dir: Path) -> pd.DataFrame:
    """Load the L2 taxonomy reference file."""
    filepath = input_dir / TAXONOMY_FILE
    logger.info(f"Loading taxonomy from {filepath}")
    df = pd.read_excel(filepath)
    logger.info(f"  Loaded {len(df)} taxonomy rows")
    return df


def load_source_data(input_dir: Path) -> pd.DataFrame:
    """Load source-item data from the most recent matching file."""
    files = sorted(input_dir.glob(SOURCE_FILE_PATTERN),
                   key=lambda f: f.stat().st_mtime)
    if not files:
        raise FileNotFoundError(
            f"No files matching '{SOURCE_FILE_PATTERN}' found in {input_dir}"
        )

    filepath = files[-1]
    logger.info(f"Loading source data from {filepath}")
    df = pd.read_excel(filepath)
    # Column name normalization — strip whitespace. If enterprise exports
    # have prefix junk (e.g. '*Event Title'), add stripping logic here.
    df.columns = [c.strip() for c in df.columns]

    required = [SOURCE_ID_COL, SOURCE_TITLE_COL, SOURCE_DESC_COL]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Source file missing required columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    pre_count = len(df)
    df[SOURCE_ID_COL] = df[SOURCE_ID_COL].astype(str).str.strip()
    df[SOURCE_TITLE_COL] = df[SOURCE_TITLE_COL].astype(str).fillna("").str.strip()
    df[SOURCE_DESC_COL] = df[SOURCE_DESC_COL].astype(str).fillna("").str.strip()
    if SOURCE_CLASS_COL in df.columns:
        df[SOURCE_CLASS_COL] = df[SOURCE_CLASS_COL].astype(str).fillna("").str.strip()
    if SOURCE_STATUS_COL in df.columns:
        df[SOURCE_STATUS_COL] = df[SOURCE_STATUS_COL].astype(str).fillna("").str.strip()

    # Exclude closed items
    if SOURCE_STATUS_COL in df.columns and CLOSED_STATUSES:
        closed_mask = df[SOURCE_STATUS_COL].str.lower().isin(CLOSED_STATUSES)
        if closed_mask.any():
            logger.info(
                f"  Excluded {closed_mask.sum()} closed items "
                f"(statuses: {df.loc[closed_mask, SOURCE_STATUS_COL].unique().tolist()})"
            )
            df = df[~closed_mask]

    # Drop rows with no meaningful text
    df = df[~((df[SOURCE_TITLE_COL].isin(["", "nan"])) &
              (df[SOURCE_DESC_COL].isin(["", "nan"])))]
    df = df[~df[SOURCE_ID_COL].isin(["", "nan"])]

    if SOURCE_ENTITY_COL in df.columns:
        df[SOURCE_ENTITY_COL] = df[SOURCE_ENTITY_COL].astype(str).str.strip()
        no_entity = df[SOURCE_ENTITY_COL].isin(["", "nan"])
        if no_entity.any():
            logger.info(f"  Dropped {no_entity.sum()} items with blank entity ID")
            df = df[~no_entity]

    logger.info(f"  Loaded {len(df)} items with text content (of {pre_count} total rows)")
    return df


def build_reference_vectors(
    nlp: "spacy.language.Language",
    tax_df: pd.DataFrame,
) -> tuple[np.ndarray, list[str], list[str]]:
    """Build one vector per unique L2 in the taxonomy.

    Aggregates by L2 name and folds L3/L4 text into each L2's semantic
    vector for richer matching. Enterprise taxonomy files are often at
    L4 grain (one row per leaf) with L2 + L2 Definition repeated across
    rows — building one vector per row would produce near-identical
    duplicates and tie the top-N ranking.

    Returns (vectors array, l2_names, l2_definitions).
    """
    def _clean(v):
        s = str(v if v is not None else "").strip()
        return "" if s.lower() in ("nan", "none") else s

    sub_cols_present = [
        c for c in (TAX_L3_COL, TAX_L3_DEF_COL, TAX_L4_COL, TAX_L4_DEF_COL)
        if c in tax_df.columns
    ]

    l2_aggregate = {}  # l2_name -> {"def": str, "text_parts": [str,...]}
    skipped = 0
    for _, row in tax_df.iterrows():
        l2_name = _clean(row.get(TAX_L2_COL))
        if not l2_name:
            skipped += 1
            continue
        if l2_name not in l2_aggregate:
            l2_aggregate[l2_name] = {"def": "", "text_parts": [l2_name]}
            l2_def = _clean(row.get(TAX_L2_DEF_COL))
            if l2_def:
                l2_aggregate[l2_name]["def"] = l2_def
                l2_aggregate[l2_name]["text_parts"].append(l2_def)
        for col in sub_cols_present:
            val = _clean(row.get(col))
            if val and val not in l2_aggregate[l2_name]["text_parts"]:
                l2_aggregate[l2_name]["text_parts"].append(val)

    if skipped:
        logger.info(f"  Skipped {skipped} taxonomy rows with blank L2")

    l2_names = list(l2_aggregate.keys())
    l2_definitions = [l2_aggregate[n]["def"] for n in l2_names]
    l2_texts = [". ".join(l2_aggregate[n]["text_parts"]) for n in l2_names]

    logger.info(
        f"Computing vectors for {len(l2_texts)} unique L2s "
        f"(aggregated from {len(tax_df)} rows)..."
    )
    vectors = [nlp(text).vector for text in l2_texts]
    vectors = np.array(vectors)
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1
    vectors = vectors / norms
    logger.info(f"  Reference vectors shape: {vectors.shape}")
    return vectors, l2_names, l2_definitions


def compute_mappings(
    nlp: "spacy.language.Language",
    source_df: pd.DataFrame,
    ref_vectors: np.ndarray,
    l2_names: list[str],
    l2_definitions: list[str],
) -> pd.DataFrame:
    """Compute semantic similarity and produce top-3 mappings per source item."""
    total = len(source_df)
    logger.info(f"Computing vectors for {total} source items...")

    results = []
    log_interval = max(1, total // 10)

    for i, (_, row) in enumerate(source_df.iterrows()):
        if i > 0 and i % log_interval == 0:
            logger.info(f"  Processed {i}/{total} ({i/total*100:.0f}%)")

        title = str(row[SOURCE_TITLE_COL])
        desc = str(row[SOURCE_DESC_COL])
        title = "" if title == "nan" else title
        desc = "" if desc == "nan" else desc
        combined = f"{title}. {desc}" if desc else title

        doc = nlp(combined)
        item_vector = doc.vector
        norm = np.linalg.norm(item_vector)
        if norm > 0:
            item_vector = item_vector / norm

        scores = ref_vectors @ item_vector
        top_indices = np.argsort(scores)[::-1][:3]
        top1, top2, top3 = top_indices[0], top_indices[1], top_indices[2]
        s1, s2, s3 = float(scores[top1]), float(scores[top2]), float(scores[top3])

        full_desc = "" if desc == "nan" else desc
        cls_val = _optional_col_value(row, SOURCE_CLASS_COL)
        status_val = _optional_col_value(row, SOURCE_STATUS_COL)

        results.append({
            "Item ID": row[SOURCE_ID_COL],
            "Entity ID": row.get(SOURCE_ENTITY_COL, ""),
            "Title": title,
            "Description": full_desc[:200],
            "Description Full": full_desc,
            "Classification": cls_val,
            "Status": status_val,
            "Match 1 - L2": l2_names[top1],
            "Match 1 - Score": round(s1, 4),
            "Match 1 - Definition": l2_definitions[top1],
            "Match 2 - L2": l2_names[top2],
            "Match 2 - Score": round(s2, 4),
            "Match 2 - Definition": l2_definitions[top2],
            "Match 3 - L2": l2_names[top3],
            "Match 3 - Score": round(s3, 4),
            "Match 3 - Definition": l2_definitions[top3],
            "Margin 1-2": round(s1 - s2, 4),
            "Margin 2-3": round(s2 - s3, 4),
            "Match 1 Valid": s1 >= MIN_SIMILARITY_SCORE,
        })

    logger.info(f"  Computed mappings for {len(results)} items")
    return pd.DataFrame(results)


def _optional_col_value(row, col_name: str) -> str:
    if col_name not in row.index:
        return ""
    val = str(row.get(col_name, ""))
    return "" if val in ("", "nan", "none") else val


def determine_ambiguity_threshold(mapping_df: pd.DataFrame) -> float:
    """Auto-tune the margin threshold from the data's P25 margin."""
    valid = mapping_df[mapping_df["Match 1 Valid"]]
    margins = valid["Margin 1-2"]
    margins = margins[margins > 0]
    if len(margins) == 0:
        return 0.02
    p25 = margins.quantile(0.25)
    median = margins.quantile(0.50)
    threshold = max(0.01, min(p25, 0.05))
    logger.info(f"  Margin distribution (valid) — P25: {p25:.4f}, median: {median:.4f}")
    logger.info(f"  Ambiguity threshold set to: {threshold:.4f}")
    return threshold


def classify_mappings(mapping_df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    """Assign Mapping Status, Match Confidence, and Mapped L2s list per row.

    - No Match: Match 1 below MIN_SIMILARITY_SCORE.
    - Needs Review: Match 1 valid but margin to Match 2 below threshold.
    - Suggested Match: Match 1 valid and margin above threshold. Additional
      L2s included if also above MIN_SIMILARITY_SCORE and within 2x the
      threshold of Match 1's score.
    """
    df = mapping_df.copy()
    statuses, confidences, l2s_list, l2_counts, defs_list = [], [], [], [], []

    for _, row in df.iterrows():
        if not row["Match 1 Valid"]:
            statuses.append("No Match")
            confidences.append("Weak")
            l2s_list.append("")
            l2_counts.append(0)
            defs_list.append("")
            continue

        margin = row["Margin 1-2"]
        if margin < threshold:
            candidates, cand_defs = [], []
            for n in (1, 2, 3):
                if row[f"Match {n} - Score"] >= MIN_SIMILARITY_SCORE:
                    candidates.append(row[f"Match {n} - L2"])
                    cand_defs.append(row[f"Match {n} - Definition"])
            statuses.append("Needs Review")
            confidences.append("Review Required")
            l2s_list.append("; ".join(candidates))
            l2_counts.append(len(candidates))
            defs_list.append("; ".join(cand_defs))
        else:
            top = row["Match 1 - Score"]
            l2s = [row["Match 1 - L2"]]
            defs = [row["Match 1 - Definition"]]
            for n in (2, 3):
                s = row[f"Match {n} - Score"]
                if s >= MIN_SIMILARITY_SCORE and (top - s) < threshold * 2:
                    l2s.append(row[f"Match {n} - L2"])
                    defs.append(row[f"Match {n} - Definition"])
            statuses.append("Suggested Match")
            confidences.append("Strong" if top >= HIGH_SIMILARITY_SCORE else "Moderate")
            l2s_list.append("; ".join(l2s))
            l2_counts.append(len(l2s))
            defs_list.append("; ".join(defs))

    df["Mapping Status"] = statuses
    df["Match Confidence"] = confidences
    df["Mapped L2s"] = l2s_list
    df["Mapped L2 Count"] = l2_counts
    df["Mapped L2 Definitions"] = defs_list
    return df


def export_results(mapping_df: pd.DataFrame, threshold: float, output_dir: Path) -> Path:
    """Write a multi-sheet Excel output with formatting.

    Sheets:
      1. All Mappings — one row per item, reviewer-friendly
      2. Needs Review — side-by-side comparison workspace
      3. Summary — counts + plain-language explanation
      4. L2 Distribution — item counts per L2 (exploded for multi-L2)
      5. Raw Scores — hidden, for development and threshold tuning
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%m%d%Y%I%M%p")
    output_path = output_dir / f"mapping_{timestamp}.xlsx"

    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    status_fills = {
        "Suggested Match": PatternFill("solid", fgColor="C6EFCE"),
        "Needs Review": PatternFill("solid", fgColor="FFFF00"),
        "No Match": PatternFill("solid", fgColor="D9D9D9"),
    }
    confidence_fills = {
        "Strong": PatternFill("solid", fgColor="C6EFCE"),
        "Moderate": PatternFill("solid", fgColor="FCE4D6"),
        "Weak": PatternFill("solid", fgColor="D9D9D9"),
        "Review Required": PatternFill("solid", fgColor="FFFF00"),
    }
    reviewer_fill = PatternFill("solid", fgColor="E2EFDA")

    def style_header(ws, max_col: int):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_fit(ws, overrides: dict | None = None, cap: int = 25):
        overrides = overrides or {}
        for col in ws.iter_cols(min_row=1, max_row=1):
            letter = get_column_letter(col[0].column)
            header = str(col[0].value or "")
            if header in overrides:
                ws.column_dimensions[letter].width = overrides[header]
            else:
                ws.column_dimensions[letter].width = min(max(len(header) + 4, 12), cap)

    def wrap_cols(ws, cols: list[str]):
        header_map = {str(c[0].value): c[0].column for c in ws.iter_cols(min_row=1, max_row=1)}
        for name in cols:
            if name in header_map:
                idx = header_map[name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).alignment = wrap_align

    def fill_col(ws, col_name: str, mapping: dict):
        header_map = {str(c[0].value): c[0].column for c in ws.iter_cols(min_row=1, max_row=1)}
        if col_name not in header_map:
            return
        idx = header_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=idx)
            f = mapping.get(str(cell.value))
            if f:
                cell.fill = f

    # Sheet 1: All Mappings
    all_cols = [
        "Item ID", "Entity ID", "Title", "Description",
        "Classification", "Status",
        "Mapping Status", "Match Confidence",
        "Mapped L2s", "Mapped L2 Count", "Mapped L2 Definitions",
    ]
    all_cols = [c for c in all_cols if c in mapping_df.columns]
    all_mappings = mapping_df[all_cols].copy()

    # Sheet 2: Needs Review
    needs_review = mapping_df[mapping_df["Mapping Status"] == "Needs Review"].copy()
    review_rows = []
    for _, r in needs_review.iterrows():
        rec = {
            "Item ID": r["Item ID"],
            "Entity ID": r["Entity ID"],
            "Title": r["Title"],
            "Description": r["Description Full"],
            "Match Confidence": r["Match Confidence"],
        }
        for n in (1, 2, 3):
            if r[f"Match {n} - Score"] >= MIN_SIMILARITY_SCORE:
                rec[f"Candidate {n} L2"] = r[f"Match {n} - L2"]
                rec[f"Candidate {n} Definition"] = r[f"Match {n} - Definition"]
            else:
                rec[f"Candidate {n} L2"] = ""
                rec[f"Candidate {n} Definition"] = ""
            rec[f"Candidate {n} Applies"] = ""
        rec["Reviewer Notes"] = ""
        review_rows.append(rec)
    review_df = pd.DataFrame(review_rows, columns=[
        "Item ID", "Entity ID", "Title", "Description", "Match Confidence",
        "Candidate 1 L2", "Candidate 1 Definition", "Candidate 1 Applies",
        "Candidate 2 L2", "Candidate 2 Definition", "Candidate 2 Applies",
        "Candidate 3 L2", "Candidate 3 Definition", "Candidate 3 Applies",
        "Reviewer Notes",
    ])

    # Sheet 3: Summary
    total = len(mapping_df)
    def pct(n):
        return f"{n} ({n/total*100:.1f}%)" if total else "0"
    suggested = (mapping_df["Mapping Status"] == "Suggested Match").sum()
    suggested_single = ((mapping_df["Mapping Status"] == "Suggested Match")
                        & (mapping_df["Mapped L2 Count"] == 1)).sum()
    suggested_multi = ((mapping_df["Mapping Status"] == "Suggested Match")
                       & (mapping_df["Mapped L2 Count"] > 1)).sum()
    needs_review_n = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()
    strong = (mapping_df["Match Confidence"] == "Strong").sum()
    moderate = (mapping_df["Match Confidence"] == "Moderate").sum()
    summary_df = pd.DataFrame({
        "Metric": [
            "Total items", "",
            "Suggested Match",
            "  to single L2", "  to multiple L2s",
            f"  Strong confidence (>= {HIGH_SIMILARITY_SCORE:.0%})",
            f"  Moderate confidence (< {HIGH_SIMILARITY_SCORE:.0%})",
            "Needs Review", "No Match",
        ],
        "Value": [
            total, "",
            pct(suggested), suggested_single, suggested_multi,
            strong, moderate,
            pct(needs_review_n), pct(no_match),
        ],
    })

    # Sheet 4: L2 Distribution (exploded on Suggested Match only)
    suggested_df = mapping_df[mapping_df["Mapping Status"] == "Suggested Match"].copy()
    exploded = suggested_df["Mapped L2s"].str.split("; ").explode().str.strip()
    exploded = exploded[exploded != ""]
    l2_dist = exploded.value_counts().reset_index()
    l2_dist.columns = ["L2", "Count (Suggested Match)"]

    # Sheet 5: Raw Scores (hidden) + stats block
    raw_cols = [
        "Item ID", "Entity ID", "Title", "Description Full",
        "Match 1 - L2", "Match 1 - Score",
        "Match 2 - L2", "Match 2 - Score",
        "Match 3 - L2", "Match 3 - Score",
        "Margin 1-2", "Margin 2-3",
        "Mapping Status", "Match Confidence", "Match 1 Valid",
    ]
    raw_scores = mapping_df[raw_cols].copy()
    raw_scores = raw_scores.rename(columns={"Description Full": "Description"})
    valid_scores = mapping_df.loc[mapping_df["Match 1 Valid"], "Match 1 - Score"]
    valid_margins = mapping_df.loc[mapping_df["Match 1 Valid"], "Margin 1-2"]
    valid_margins = valid_margins[valid_margins > 0]
    raw_stats_df = pd.DataFrame({
        "Metric": [
            "Score distribution (valid Match 1)",
            "  Mean", "  Median", "  Min", "  Max", "",
            "Margin distribution (valid, non-zero)",
            "  P25", "  P50", "  P75", "",
            "Settings",
            "  Ambiguity threshold",
            "  Min similarity score",
            "  spaCy model",
        ],
        "Value": [
            "",
            f"{valid_scores.mean():.4f}" if len(valid_scores) else "N/A",
            f"{valid_scores.median():.4f}" if len(valid_scores) else "N/A",
            f"{valid_scores.min():.4f}" if len(valid_scores) else "N/A",
            f"{valid_scores.max():.4f}" if len(valid_scores) else "N/A", "",
            "",
            f"{valid_margins.quantile(0.25):.4f}" if len(valid_margins) else "N/A",
            f"{valid_margins.quantile(0.50):.4f}" if len(valid_margins) else "N/A",
            f"{valid_margins.quantile(0.75):.4f}" if len(valid_margins) else "N/A", "",
            "",
            f"{threshold:.4f}",
            MIN_SIMILARITY_SCORE,
            SPACY_MODEL,
        ],
    })

    logger.info(f"Writing output to {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_mappings.to_excel(writer, sheet_name="All Mappings", index=False)
        review_df.to_excel(writer, sheet_name="Needs Review", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        l2_dist.to_excel(writer, sheet_name="L2 Distribution", index=False)
        raw_scores.to_excel(writer, sheet_name="Raw Scores", index=False)

        wb = writer.book

        ws = wb["All Mappings"]
        style_header(ws, ws.max_column)
        auto_fit(ws, overrides={
            "Description": 60, "Title": 30,
            "Mapped L2s": 50, "Mapped L2 Definitions": 60,
        })
        wrap_cols(ws, ["Description", "Mapped L2s", "Mapped L2 Definitions"])
        fill_col(ws, "Mapping Status", status_fills)
        fill_col(ws, "Match Confidence", confidence_fills)
        ws.freeze_panes = "C2"

        ws = wb["Needs Review"]
        style_header(ws, ws.max_column)
        auto_fit(ws, overrides={
            "Description": 60, "Title": 30,
            "Candidate 1 Definition": 60, "Candidate 2 Definition": 60,
            "Candidate 3 Definition": 60,
            "Candidate 1 L2": 25, "Candidate 2 L2": 25, "Candidate 3 L2": 25,
            "Candidate 1 Applies": 15, "Candidate 2 Applies": 15,
            "Candidate 3 Applies": 15, "Reviewer Notes": 30,
        })
        wrap_cols(ws, [
            "Description",
            "Candidate 1 Definition", "Candidate 2 Definition", "Candidate 3 Definition",
        ])
        fill_col(ws, "Match Confidence", confidence_fills)
        ws.freeze_panes = "A2"
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 60
        reviewer_headers = {
            "Candidate 1 Applies", "Candidate 2 Applies", "Candidate 3 Applies",
            "Reviewer Notes",
        }
        for col in ws.iter_cols(min_row=1, max_row=1):
            if str(col[0].value) in reviewer_headers:
                col[0].fill = reviewer_fill

        ws = wb["Summary"]
        style_header(ws, ws.max_column)
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 25

        ws = wb["L2 Distribution"]
        style_header(ws, ws.max_column)
        auto_fit(ws, overrides={"L2": 45})

        ws = wb["Raw Scores"]
        style_header(ws, ws.max_column)
        auto_fit(ws, overrides={"Description": 60, "Title": 30})
        wrap_cols(ws, ["Description"])
        stats_start = ws.max_row + 3
        for i, r in raw_stats_df.iterrows():
            ws.cell(row=stats_start + i, column=1, value=r["Metric"])
            ws.cell(row=stats_start + i, column=2, value=r["Value"])
        ws.sheet_state = "hidden"

    logger.info(f"  Output saved: {output_path}")
    return output_path


# =============================================================================
# MAIN
# =============================================================================

def main():
    global AMBIGUITY_MARGIN_THRESHOLD

    input_dir = _PROJECT_ROOT / "data" / "input"
    output_dir = _PROJECT_ROOT / "data" / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    tax_df = load_taxonomy(input_dir)
    source_df = load_source_data(input_dir)

    logger.info(f"Loading spaCy model: {SPACY_MODEL}")
    nlp = spacy.load(SPACY_MODEL)
    logger.info(
        f"  Model loaded ({len(nlp.vocab.vectors)} vectors, "
        f"{nlp.vocab.vectors.shape[1]} dimensions)"
    )

    ref_vectors, l2_names, l2_definitions = build_reference_vectors(nlp, tax_df)
    mapping_df = compute_mappings(nlp, source_df, ref_vectors, l2_names, l2_definitions)

    if AMBIGUITY_MARGIN_THRESHOLD is None:
        AMBIGUITY_MARGIN_THRESHOLD = determine_ambiguity_threshold(mapping_df)

    mapping_df = classify_mappings(mapping_df, AMBIGUITY_MARGIN_THRESHOLD)

    total = len(mapping_df)
    suggested = (mapping_df["Mapping Status"] == "Suggested Match").sum()
    single = ((mapping_df["Mapping Status"] == "Suggested Match")
              & (mapping_df["Mapped L2 Count"] == 1)).sum()
    multi = ((mapping_df["Mapping Status"] == "Suggested Match")
             & (mapping_df["Mapped L2 Count"] > 1)).sum()
    needs_review = (mapping_df["Mapping Status"] == "Needs Review").sum()
    no_match = (mapping_df["Mapping Status"] == "No Match").sum()

    logger.info("=" * 60)
    logger.info("MAPPING COMPLETE")
    logger.info(f"  Total items: {total}")
    if total:
        logger.info(f"  Suggested Match: {suggested} ({suggested/total*100:.1f}%) — single: {single}, multi: {multi}")
        logger.info(f"  Needs Review: {needs_review} ({needs_review/total*100:.1f}%)")
        logger.info(f"  No Match: {no_match} ({no_match/total*100:.1f}%)")
    logger.info(f"  Ambiguity threshold: {AMBIGUITY_MARGIN_THRESHOLD:.4f}")
    logger.info("=" * 60)

    output_path = export_results(mapping_df, AMBIGUITY_MARGIN_THRESHOLD, output_dir)

    print(f"\nDone! Output: {output_path}")
    print(f"  Suggested Match: {suggested} (single: {single}, multi: {multi}) "
          f"| Needs Review: {needs_review} | No Match: {no_match}")


if __name__ == "__main__":
    main()
